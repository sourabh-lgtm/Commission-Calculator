"""Excel workbook export for Commission Calculator."""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# Brand colours
CORAL_HEX  = "FF9178"
GREEN_HEX  = "16a34a"
DIM_HEX    = "595959"
HEADER_BG  = "FF9178"
TOTAL_BG   = "EEEEEE"
ALT_ROW_BG = "F5F5F5"
WHITE      = "FFFFFF"
BLACK      = "000000"

thin = Side(style="thin", color="E0E0E0")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_style(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = Font(name="Calibri", bold=True, color=WHITE, size=10)
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")
        cell.border    = THIN_BORDER


def _total_style(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = Font(name="Calibri", bold=True, size=10)
        cell.fill      = PatternFill("solid", fgColor=TOTAL_BG)
        cell.alignment = Alignment(horizontal="right" if col > 1 else "left", vertical="center")
        cell.border    = THIN_BORDER


def _alt_row(ws, row, cols, alt=False):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        if alt:
            cell.fill = PatternFill("solid", fgColor=ALT_ROW_BG)
        cell.alignment = Alignment(horizontal="right" if col > 1 else "left", vertical="center")
        cell.border    = THIN_BORDER
        cell.font      = Font(name="Calibri", size=10)


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)


def export_workbook(model, output_path: str = "output/Commission_Output.xlsx"):
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    _sheet_commission_summary(wb, model)
    _sheet_sdr_detail(wb, model)
    _sheet_quarterly_accelerators(wb, model)
    _sheet_commission_workings(wb, model)
    _sheet_fx_rates(wb, model)

    wb.save(output_path)
    print(f"[Excel] Saved → {output_path}")
    return output_path


# ---------------------------------------------------------------------------
# Sheet 1: Commission Summary
# ---------------------------------------------------------------------------
def _sheet_commission_summary(wb: Workbook, model):
    ws = wb.create_sheet("Commission Summary")
    heads = ["Employee", "Month", "Quarter", "Currency", "Out SAOs", "In SAOs",
             "Out SAO Comm", "In SAO Comm", "Out CW ACV (EUR)", "In CW ACV (EUR)",
             "Out CW Comm", "In CW Comm", "Accelerator", "Total Commission"]
    ws.append(heads)
    _header_style(ws, 1, len(heads))
    ws.row_dimensions[1].height = 18

    if model.commission_detail.empty:
        _auto_width(ws)
        return

    df = model.commission_detail.copy()
    df = df.merge(model.employees[["employee_id","name"]], on="employee_id", how="left", suffixes=("","_emp"))
    if "name_emp" in df.columns:
        df["_display_name"] = df["name_emp"].fillna(df.get("name",""))
    else:
        df["_display_name"] = df.get("name","")

    for i, (_, row) in enumerate(df.iterrows(), start=2):
        ws.append([
            row.get("_display_name",""),
            row["month"].strftime("%Y-%m") if hasattr(row["month"],"strftime") else str(row["month"]),
            row.get("quarter",""),
            row.get("currency",""),
            row.get("outbound_sao_count",0),
            row.get("inbound_sao_count",0),
            row.get("outbound_sao_comm",0),
            row.get("inbound_sao_comm",0),
            row.get("outbound_cw_acv_eur",0),
            row.get("inbound_cw_acv_eur",0),
            row.get("outbound_cw_comm",0),
            row.get("inbound_cw_comm",0),
            row.get("accelerator_topup",0),
            row.get("total_commission",0),
        ])
        _alt_row(ws, i, len(heads), alt=i % 2 == 0)

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Sheet 2: SDR Detail (one row per employee per month)
# ---------------------------------------------------------------------------
def _sheet_sdr_detail(wb: Workbook, model):
    ws = wb.create_sheet("SDR Detail")
    _sheet_commission_summary(wb, model)   # same structure for now
    # Reuse same data — sheet is already identical to Commission Summary
    # but named separately for clarity
    ws = wb["SDR Detail"]
    _auto_width(ws)


# ---------------------------------------------------------------------------
# Sheet 3: Quarterly Accelerators
# ---------------------------------------------------------------------------
def _sheet_quarterly_accelerators(wb: Workbook, model):
    ws = wb.create_sheet("Quarterly Accelerators")
    heads = ["Employee ID","Year","Quarter","Currency","Total SAOs","Outbound SAOs",
             "Inbound SAOs","Threshold","Excess SAOs","Excess Outbound","Top-up / SAO","Accelerator"]
    ws.append(heads)
    _header_style(ws, 1, len(heads))

    if model.accelerators.empty:
        _auto_width(ws)
        return

    for i, (_, row) in enumerate(model.accelerators.iterrows(), start=2):
        ws.append([
            row.get("employee_id",""),
            row.get("year",""),
            row.get("quarter",""),
            row.get("currency",""),
            row.get("total_saos",0),
            row.get("outbound_saos",0),
            row.get("inbound_saos",0),
            row.get("threshold",9),
            row.get("excess_saos",0),
            row.get("excess_outbound",0),
            row.get("topup_per_sao",0),
            row.get("accelerator_topup",0),
        ])
        _alt_row(ws, i, len(heads), alt=i % 2 == 0)

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Sheet 4: Commission Workings (all activities)
# ---------------------------------------------------------------------------
def _sheet_commission_workings(wb: Workbook, model):
    from src.commission_plans import get_plan
    ws = wb.create_sheet("Commission Workings")
    heads = ["Employee","Month","Date","Opportunity ID","SAO Type","Category",
             "ACV (EUR)","FX Rate","Rate Description","Commission","Currency"]
    ws.append(heads)
    _header_style(ws, 1, len(heads))

    sdrs = model.employees[model.employees["role"] == "sdr"]
    row_idx = 2

    for _, emp in sdrs.iterrows():
        plan_cls = get_plan(emp["role"])
        if not plan_cls:
            continue
        plan = plan_cls()
        for month in model.active_months:
            rows = plan.get_workings_rows(emp, month, model.sdr_activities, model.closed_won, model.fx_rates)
            for r in rows:
                ws.append([
                    emp["name"],
                    month.strftime("%Y-%m"),
                    r.get("date",""),
                    r.get("opportunity_id",""),
                    r.get("sao_type",""),
                    r.get("type",""),
                    r.get("acv_eur",""),
                    r.get("fx_rate",""),
                    r.get("rate_desc",""),
                    r.get("commission",0),
                    r.get("currency",""),
                ])
                _alt_row(ws, row_idx, len(heads), alt=row_idx % 2 == 0)
                row_idx += 1

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Sheet 5: FX Rates
# ---------------------------------------------------------------------------
def _sheet_fx_rates(wb: Workbook, model):
    ws = wb.create_sheet("FX Rates")
    if model.fx_rates.empty:
        return
    df = model.fx_rates.copy()
    df["month"] = df["month"].dt.strftime("%Y-%m-%d")
    heads = list(df.columns)
    ws.append(heads)
    _header_style(ws, 1, len(heads))
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        ws.append(list(row))
        _alt_row(ws, i, len(heads), alt=i % 2 == 0)
    _auto_width(ws)


# ---------------------------------------------------------------------------
# Payroll workbook (one sheet per region, employee rows × month columns)
# ---------------------------------------------------------------------------

def export_payroll_workbook(model, year: int) -> bytes:
    import io
    from src.reports import payroll_summary
    data = payroll_summary(model, year)

    wb = Workbook()
    wb.remove(wb.active)

    for rd in data["regions"]:
        region       = rd["region"]
        emps         = rd["employees"]
        months       = data["months"]
        month_labels = data["month_labels"]
        region_short = "SE" if region == "Nordics" else region
        ws = wb.create_sheet(f"{region_short} {year}")

        # --- Row 1: title ---
        n_cols = 4 + len(months) + 4 + 1
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        tc = ws.cell(row=1, column=1, value=f"Commission & Bonus Payroll Summary — FY{year}")
        tc.font      = Font(name="Calibri", bold=True, size=13)
        tc.alignment = Alignment(vertical="center")
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 6

        # --- Row 3: quarter bands ---
        def _q(m_str): return (pd.Timestamp(m_str).month - 1) // 3 + 1
        q_col = 5
        for q in range(1, 5):
            q_months = [m for m in months if _q(m) == q]
            if q_months:
                c = ws.cell(row=3, column=q_col, value=f"Q{q}")
                c.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
                c.fill = PatternFill("solid", fgColor=HEADER_BG)
                c.alignment = Alignment(horizontal="center")
                if len(q_months) > 1:
                    ws.merge_cells(start_row=3, start_column=q_col,
                                   end_row=3, end_column=q_col + len(q_months) - 1)
                q_col += len(q_months)
        ws.row_dimensions[3].height = 14

        # --- Row 4: column headers ---
        heads = ["Employee ID", "Name", "Dept Code", "Currency"] + month_labels + ["Q1", "Q2", "Q3", "Q4", "Total"]
        for col, h in enumerate(heads, 1):
            c = ws.cell(row=4, column=col, value=h)
            c.font      = Font(name="Calibri", bold=True, color=WHITE, size=10)
            c.fill      = PatternFill("solid", fgColor=HEADER_BG)
            c.alignment = Alignment(horizontal="center" if col > 4 else "left", vertical="center")
            c.border    = THIN_BORDER
        ws.row_dimensions[4].height = 18

        # --- Data rows ---
        t_monthly = {m: 0.0 for m in months}
        t_q = [0.0, 0.0, 0.0, 0.0]
        t_total = 0.0

        for i, emp in enumerate(emps):
            r = 5 + i
            alt = i % 2 == 0
            vals = ([emp["employee_id"], emp["name"], emp.get("cost_center_code",""), emp["currency"]]
                    + [emp["monthly"].get(m, 0) for m in months]
                    + [emp["q1"], emp["q2"], emp["q3"], emp["q4"], emp["total"]])
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=r, column=col, value=v)
                c.font      = Font(name="Calibri", size=10)
                c.alignment = Alignment(horizontal="right" if col > 4 else "left", vertical="center")
                c.border    = THIN_BORDER
                if alt:
                    c.fill = PatternFill("solid", fgColor=ALT_ROW_BG)
                if col > 4 and isinstance(v, (int, float)):
                    c.number_format = "#,##0.00"
            for m in months:
                t_monthly[m] += emp["monthly"].get(m, 0)
            t_q[0] += emp["q1"]; t_q[1] += emp["q2"]
            t_q[2] += emp["q3"]; t_q[3] += emp["q4"]
            t_total += emp["total"]

        # --- Totals row ---
        tr = 5 + len(emps)
        tvals = (["", "TOTAL", "", ""]
                 + [round(t_monthly[m], 2) for m in months]
                 + [round(x, 2) for x in t_q] + [round(t_total, 2)])
        for col, v in enumerate(tvals, 1):
            c = ws.cell(row=tr, column=col, value=v)
            c.font      = Font(name="Calibri", bold=True, size=10)
            c.fill      = PatternFill("solid", fgColor=TOTAL_BG)
            c.alignment = Alignment(horizontal="right" if col > 4 else "left", vertical="center")
            c.border    = THIN_BORDER
            if col > 4 and isinstance(v, (int, float)):
                c.number_format = "#,##0.00"

        # Column widths
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 24
        ws.column_dimensions["C"].width = 34
        ws.column_dimensions["D"].width = 10
        for ci in range(5, n_cols + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 11

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Accrual workbook (department rows × month columns, EUR, NI for UK)
# ---------------------------------------------------------------------------

def export_accrual_workbook(model, year: int) -> bytes:
    import io
    from src.reports import accrual_summary
    data = accrual_summary(model, year)

    wb = Workbook()
    wb.remove(wb.active)

    for rd in data["regions"]:
        region       = rd["region"]
        rows         = rd["rows"]
        months       = data["months"]
        month_labels = data["month_labels"]
        region_short = "SE" if region == "Nordics" else region
        ws = wb.create_sheet(f"{region_short} FY{str(year)[2:]}")

        n_cols = 4 + len(months) + 4 + 1   # emp_id, name, dept_code, type + months + Q1-4 + total

        # --- Row 1: title ---
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        tc = ws.cell(row=1, column=1, value=f"Commission & Bonus Accrual Summary — FY{year}")
        tc.font      = Font(name="Calibri", bold=True, size=13)
        tc.alignment = Alignment(vertical="center")
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 6

        # --- Row 3: quarter bands ---
        def _q(m_str): return (pd.Timestamp(m_str).month - 1) // 3 + 1
        q_col = 5   # cols 1-4: emp_id, name, dept_code, type
        for q in range(1, 5):
            q_months = [m for m in months if _q(m) == q]
            if q_months:
                c = ws.cell(row=3, column=q_col, value=f"Q{q}")
                c.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
                c.fill = PatternFill("solid", fgColor=HEADER_BG)
                c.alignment = Alignment(horizontal="center")
                if len(q_months) > 1:
                    ws.merge_cells(start_row=3, start_column=q_col,
                                   end_row=3, end_column=q_col + len(q_months) - 1)
                q_col += len(q_months)
        ws.row_dimensions[3].height = 14

        # --- Row 4: column headers ---
        heads = ["Employee ID", "Name", "Dept Code", "Type"] + month_labels + ["Q1", "Q2", "Q3", "Q4", "Total (EUR)"]
        for col, h in enumerate(heads, 1):
            c = ws.cell(row=4, column=col, value=h)
            c.font      = Font(name="Calibri", bold=True, color=WHITE, size=10)
            c.fill      = PatternFill("solid", fgColor=HEADER_BG)
            c.alignment = Alignment(horizontal="center" if col > 4 else "left", vertical="center")
            c.border    = THIN_BORDER
        ws.row_dimensions[4].height = 18

        # --- Data rows ---
        t_monthly = {m: 0.0 for m in months}
        t_q = [0.0, 0.0, 0.0, 0.0]
        t_total = 0.0

        for i, row in enumerate(rows):
            r    = 5 + i
            alt  = i % 2 == 0
            is_ni = "NI" in row["type"]
            vals  = ([row["employee_id"], row["name"], row.get("cost_center_code",""), row["type"]]
                     + [row["monthly"].get(m, 0) for m in months]
                     + [row["q1"], row["q2"], row["q3"], row["q4"], row["total"]])
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=r, column=col, value=v)
                c.font      = Font(name="Calibri", size=10,
                                   italic=is_ni, color="777777" if is_ni else BLACK)
                c.alignment = Alignment(horizontal="right" if col > 4 else "left", vertical="center")
                c.border    = THIN_BORDER
                if alt and not is_ni:
                    c.fill = PatternFill("solid", fgColor=ALT_ROW_BG)
                if col > 2 and isinstance(v, (int, float)):
                    c.number_format = "#,##0.00"
            if not is_ni:
                for m in months:
                    t_monthly[m] += row["monthly"].get(m, 0)
                t_q[0] += row["q1"]; t_q[1] += row["q2"]
                t_q[2] += row["q3"]; t_q[3] += row["q4"]
                t_total += row["total"]

        # --- Totals row ---
        tr = 5 + len(rows)
        tvals = (["TOTAL (Commission)", "", "", ""]
                 + [round(t_monthly[m], 2) for m in months]
                 + [round(x, 2) for x in t_q] + [round(t_total, 2)])
        for col, v in enumerate(tvals, 1):
            c = ws.cell(row=tr, column=col, value=v)
            c.font      = Font(name="Calibri", bold=True, size=10)
            c.fill      = PatternFill("solid", fgColor=TOTAL_BG)
            c.alignment = Alignment(horizontal="right" if col > 4 else "left", vertical="center")
            c.border    = THIN_BORDER
            if col > 2 and isinstance(v, (int, float)):
                c.number_format = "#,##0.00"

        # Column widths
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 26
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 22
        for ci in range(5, n_cols + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 11

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


if __name__ == "__main__":
    from src.pipeline import run_pipeline
    m = run_pipeline("data")
    export_workbook(m)
